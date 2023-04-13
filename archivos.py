from os import getcwd
import os 
import openpyxl as op
import sqlite3 as sql
from tabulate import tabulate
import pandas as pd

class APP:
    pass



class archivo:
    def __init__(self,nombre) -> None:
        self.ruta = getcwd().replace("\\","/")
        self.nombre = nombre
        self.wb = self.leer_archivo()     #leer arachi                           
        if self.wb:
            self.ws = self.wb.active
            self.last_row = self.ws.max_row +1

    def leer_archivo(self):  #RF1 ,RF3 ,RF5
        try:
           existe =  op.load_workbook(self.ruta + self.nombre) 
           if existe:
                print(f'se cargo correctamente {self.nombre}')
                return existe
        except Exception:   
            print(f'no existe el archivo {self.nombre}')
            return None
        
    def cargar_archivo(self):
        pass    


class proyecto:  
    def __init__(self,ref,cuenta,fecha_inicio) -> None:
         self.ref = ref
         self.cuenta = cuenta
         self.fecha_inicio = fecha_inicio


class archivo_proyecto(archivo):
    def __init__(self, nombre) -> None:
        super().__init__(nombre)
        self.proyectos = dict()
        if self.wb:
            self.cargar_archivo()

    def imprimir_proyectos(self):
            self.proyectos = sorted(self.proyectos.items() )
            for i in self.proyectos:
                print(i)    

    def cargar_archivo(self):                  #RF4 
        for row in range(2,self.last_row,1):
            cuenta = self.ws.cell(row,3).value       
            if cuenta not in self.proyectos:
                ref = self.ws.cell(row,2).value
                fecha = self.ws.cell(row,4).value.strftime('%Y-%m-%d')
                #instacia proyectos
                proyec = proyecto(ref, cuenta, fecha )
                self.proyectos[cuenta] = proyec


class archivo_seven(archivo):
    def __init__(self, nombre) -> None:
        super().__init__(nombre)
        self.registros =[]
        self.registros_error =[]
        self.proyectos: dict = archivo_proyecto('/proyectos.xlsx')

        if self.wb:
            print(self.cargar_archivo())

    def imprimir_seven (self):
            for i in self.registros:
                print(i)

    def imprimir_errores(self):
            for i in self.registros_error:
                print(i)            

    def cargar_archivo(self):       #RF2     
        try:
            for row in range(2,self.last_row):
                celda1 = self.ws.cell(row,1).value
                debito = self.ws.cell(row,7).value
                if  celda1 == 'PROYECTO':            #RF7 --> Verificar proyecto
                    cuenta = str(self.ws.cell(row,2).value)        

                if isinstance(debito,(float,int)) and  debito > 0:   #RF8 -- verificar debito 
                    ref = self.ws.cell(row,4).value
                    if str(cuenta) in self.proyectos.proyectos:
                        fec_registro = celda1.strftime('%Y-%m-%d')
                        fec_inicio = self.proyectos.proyectos[str(cuenta)].fecha_inicio
                        self.registros.append((row,fec_registro,fec_inicio,ref,cuenta,debito))
                    else:
                        self.registros_error.append((row,ref,cuenta,debito))   
    
            return len(self.registros ) + len(self.registros_error) #self.proyectos.print()   
        except:            
            return row ,self.ws.cell(row,7).value     


class archivo_pure(archivo):  
    def __init__(self, nombre) -> None:
        super().__init__(nombre)
        self.registros=[]
        if self.wb:
            self.cargar_archivo()
            self.crear_BD()
    
     
    def imprimir_pure(self,lista):
        for i in lista:  
            print(i)
      
    def cargar_archivo(self): #RF2   
        self.registros=[]
        for fila in self.ws.iter_rows(min_row=2,max_col=14,values_only=True):
            self.registros.append(list(fila))      

    
    def confirmar_BD(self):
        if os.path.isfile(self.ruta +'/pure.db'):
            print(f'Base de datos existe')
            return True
        print(f'Base de datos NO existe')
          
    def crear_BD(self): 
        self.conexion = sql.connect('pure.db')
        self.cursor = self.conexion.cursor()
        self.cursor.execute(""" CREATE TABLE IF NOT EXISTS pure (
        id TEXT NOT NULL,
        type TEXT NOT NULL,
        title TEXT NOT NULL,
        managedByOrganisation TEXT NOT NULL,
        awardDate TEXT NOT NULL,
        idFunding TEXT NOT NULL,
        idBudget TEXT NOT NULL,
        budgetLine INTEGER NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        expenditureValue INTEGER NOT NULL,
        visibility TEXT NOT NULL,
        observaciones TEXT ,
        usuario TEXT 
    )""")
        self.llenar_db()
    
    def llenar_db(self):
        try:
            c=1
            a=[]
            for registro in self.registros:
                c+=1 
                self.cursor.execute("""SELECT * FROM pure 
                                    WHERE id = ? AND type = ? AND title = ? AND  managedByOrganisation = ? AND
                                    awardDate = ?  AND  idFunding = ? AND idBudget = ? AND budgetLine = ? AND
                                    year = ? AND month = ? AND expenditureValue = ? AND visibility = ? AND
                                    observaciones = ? AND usuario = ? """   , (str(registro[0]) , str(registro[1]) , str(registro[2]) , str(registro[3]) ,str(registro[4]) , str(registro[5]),str(registro[6]) , str(registro[7]) , int(registro[8]) , int(registro[9]) ,int(registro[10]) , str(registro[11]), str(registro[12]) , str(registro[13])))
                if  self.cursor.fetchone() == None:
                    self.conexion.execute("INSERT INTO pure VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ? ,?)",(str(registro[0]) , str(registro[1]) , str(registro[2]) , str(registro[3]) ,str(registro[4]) , str(registro[5]),str(registro[6]) , str(registro[7]) , int(registro[8]) , int(registro[9]) ,int(registro[10]) , str(registro[11]), str(registro[12]) , str(registro[13])))
                else:
                    a.append(c)                          

        except:
            print('hubo error al cargar archivo pure')  
        self.conexion.commit()
        self.conexion.close()   
            


# seven = archivo_seven('/gastos_seven.xlsx')
# conn = sql.connect('Pure.db')
# print(conn)
# pure = archivo_pure('/gastos_pure.xlsx') 
# pure.imprimir_pure(pure.registros)

# Proy=archivo_proyecto('/proyectos.xlsx')
# Proy.imprimir_proyectos()
# seven.print()









# if __name__ == "__main__":

#     def menu():
#         print('            Migrar de SEVEN a PURE:')
#         print('Menú')
#         print('')
      
#         print('cargar archivo seven: (1)')
#         print('cargar archivo pure: (2)')
#         print('cargar archivo proyecto: (3)')

#         print('salir (0)')
#         print('----------------------------------------')

#     rta = 9999
#     print(isinstance(rta,int), 'este es')
#     menu()
#     while rta != 0:
#         rta = int(input('seleccione una del las opciones anteriores: '))
#         if isinstance(rta,int):
#             if 0< rta and rta <= 2 :
#                 if rta == 1 :
#                     nombre_seven = input('Ingrese nombre del archivo seven: ')
#                     seven = archivo_seven("/" + nombre_seven)

#                 elif rta ==2 :
#                     nombre_pure= input('Ingrese nombre del archivo pure: ')
#                     seven = archivo_pure("/" + nombre_pure) 
#             else:     
#                 print('opcion no valida')
#         else:  
#             print('opcion no valida, no es un entero')      

#         rta = 999
#         print('----------------------------------------')   
#         print('')  
#         menu()

#     print('cerro app: ')    