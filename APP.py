from os import getcwd
import os 
import openpyxl as op
import sqlite3 as sql
from tabulate import tabulate
import pandas as pd
import datetime


class Usuario:
    pass



class Archivo:
    def __init__(self,nombre) -> None:
        self.ruta = getcwd().replace("\\","/")
        self.nombre = nombre
        self.wb = self.leer_archivo()     #leer arachi                           
        if self.wb:
            self.ws = self.wb.active
            self.last_row = self.ws.max_row +1

    def leer_archivo(self):  #RF1 ,RF3 ,RF5
        try:
           existe =  op.load_workbook(self.ruta + self.nombre) 
           if existe:
                print(f'se cargo correctamente {self.nombre}')
                return existe
        except Exception:   
            print(f'no existe el archivo {self.nombre}')
            return None
        
    def cargar_archivo(self):
        pass    


class Proyecto:  
    def __init__(self,id,cuenta,fecha_inicio) -> None:
         self.id = id
         self.cuenta = cuenta
         self.fecha_inicio = fecha_inicio


class Archivo_proyecto(Archivo):
    def __init__(self, nombre) -> None:
        super().__init__(nombre)
        self.proyectos = dict()
        if self.wb:
            self.cargar_archivo()

    def imprimir_proyectos(self):
            self.proyectos = sorted(self.proyectos.items() )
            for i in self.proyectos:
                print(i)    

    def cargar_archivo(self):                  #RF4 
        for row in range(2,self.last_row,1):
            cuenta = self.ws.cell(row,3).value       
            if cuenta not in self.proyectos:
                id = self.ws.cell(row,2).value
                fecha = self.ws.cell(row,4).value.strftime('%Y-%m-%d')
                #instacia proyectos
                proyec = Proyecto(id, cuenta, fecha )
                self.proyectos[cuenta] = proyec


class Archivo_seven(Archivo):
    def __init__(self, nombre) -> None:
        super().__init__(nombre)
        self.registros =[]
        self.registros_error =[]
        self.proyectos: dict = Archivo_proyecto('/proyectos.xlsx')

        if self.wb:
            self.cargar_archivo()

    def imprimir_seven(self):
            for i in self.registros:
                print(i)

    def imprimir_errores(self):
            for i in self.registros_error:
                print(i)            

    def cargar_archivo(self):       #RF2     
        TYPE="researchcouncils/award"
        TITLE = "dummy"
        MANAGED_BY_ORG = "org"
        VISIBILITY = "confidencial"
        USER ="Faculty_of_Engineering"

        try:
            for row in range(2,self.last_row):
                celda1 = self.ws.cell(row,1).value
                debito = self.ws.cell(row,7).value
                if  celda1 == 'PROYECTO':            #RF7 --> Verificar proyecto
                    cuenta = str(self.buscar_codigo_contable(row))    

                if isinstance(debito,(float,int)) and  debito > 0:   #RF8 -- verificar debito 
                    id = self.ws.cell(row,4).value
                    if str(cuenta) in self.proyectos.proyectos:
                        #capturar datos
                        #variables
                        fec_registro = self.buscar_fecha_registro(celda1)
                        observacion = self.buscar_observacion(row)
                        ano_gasto , mes_gasto = self.separar_fecha(fec_registro)
                        fec_inicio = self.proyectos.proyectos[str(cuenta)].fecha_inicio
                        self.registros.append((row,id,TYPE,TITLE,MANAGED_BY_ORG,fec_inicio,"","",cuenta,ano_gasto,mes_gasto,debito,VISIBILITY,observacion,USER))
                    else:
                        self.registros_error.append((row,id,cuenta,debito))   
    
            return len(self.registros ) + len(self.registros_error) #self.proyectos.print()   
        except:            
            return row ,self.ws.cell(row,7).value   
        

    def buscar_fecha_registro(self,celda1):
        return celda1.strftime('%Y-%m-%d')#RF10 -- buscar fecha
    
    def separar_fecha(self,fec_registro):
        ano_gasto = datetime.datetime.strptime(fec_registro, '%Y-%m-%d').year
        mes_gasto = datetime.datetime.strptime(fec_registro, '%Y-%m-%d').month
        return ano_gasto, mes_gasto
    
    def buscar_observacion(self,row):
        return self.ws.cell(row,5).value #RF10 -- buscar Observación
    
    def buscar_codigo_contable(self,row):
        return self.ws.cell(row,2).value #RF11 -- buscar cuenta contable



    def llevar_seven_a_pure(self):
        pass

class Archivo_pure(Archivo):  
    def __init__(self, nombre) -> None:
        super().__init__(nombre)
        self.registros=[]
        if self.wb:
            self.cargar_archivo()
            self.crear_BD()
            self.llenar_DB_inicial()
    
     
    def imprimir_pure(self):
        for i in self.registros:  
            print(i)
      
    def cargar_archivo(self): #RF2   
        self.registros=[]
        for fila in self.ws.iter_rows(min_row=2,max_col=14,values_only=True):
            self.registros.append(list(fila))      
   
    def crear_BD(self): 
        self.conexion = sql.connect('pure.db')
        self.cursor = self.conexion.cursor()
        self.cursor.execute(""" CREATE TABLE IF NOT EXISTS pure (
        id TEXT NOT NULL,
        type TEXT NOT NULL,
        title TEXT NOT NULL,
        managedByOrganisation TEXT NOT NULL,
        awardDate TEXT NOT NULL,
        idFunding TEXT ,
        idBudget TEXT ,
        budgetLine INTEGER NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        expenditureValue INTEGER NOT NULL,
        visibility TEXT NOT NULL,
        observaciones TEXT ,
        usuario TEXT 
    )""")
    
    def llenar_DB_inicial(self):
        try:
            c=1
            a=[]
            for registro in self.registros:
                c+=1 
                self.cursor.execute("""SELECT * FROM pure 
                                    WHERE id = ? AND type = ? AND title = ? AND  managedByOrganisation = ? AND
                                    awardDate = ?  AND  idFunding = ? AND idBudget = ? AND budgetLine = ? AND
                                    year = ? AND month = ? AND expenditureValue = ? AND visibility = ? AND
                                    observaciones = ? AND usuario = ? """ , (str(registro[0]) , str(registro[1]) , str(registro[2]) , str(registro[3]) ,str(registro[4]) , str(registro[5]),str(registro[6]) , str(registro[7]) , int(registro[8]) , int(registro[9]) ,int(registro[10]) , str(registro[11]), str(registro[12]) , str(registro[13])))
                if  self.cursor.fetchone() == None:
                    self.conexion.execute("INSERT INTO pure VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ? ,?)",(str(registro[0]) , str(registro[1]) , str(registro[2]) , str(registro[3]) ,str(registro[4]) , str(registro[5]),str(registro[6]) , str(registro[7]) , int(registro[8]) , int(registro[9]) ,int(registro[10]) , str(registro[11]), str(registro[12]) , str(registro[13])))
                else:
                    a.append(c)                          

        except:
            print('hubo error al cargar archivo pure')  
        self.conexion.commit()
        self.conexion.close()   
            
    def verificar_registros(self,seven:Archivo_seven):
        num=0
        self.crear_BD()
        for registro in seven.registros:
            self.cursor.execute("""SELECT * FROM pure 
                                    WHERE id = ? AND type = ? AND title = ? AND  managedByOrganisation = ? AND
                                    awardDate = ?  AND  idFunding = ? AND idBudget = ? AND budgetLine = ? AND
                                    year = ? AND month = ? AND expenditureValue = ? AND visibility = ? AND
                                    observaciones = ? AND usuario = ? """ , (str(registro[1]) , str(registro[2]) , str(registro[3]) , str(registro[4]) ,str(registro[5]) , str(registro[6]),str(registro[7]) , str(registro[8]) , int(registro[9]) , int(registro[10]) ,int(registro[11]) , str(registro[12]), str(registro[13]) , str(registro[14])))
            if  self.cursor.fetchone() == None:
                self.registrar_seven(registro)
                num+=1

        self.conexion.commit()           
        self.conexion.close() 
        print(f'\n----------------------------------------{num}') 
        if num>0 :
            print(f'{num} gastos registrados') 
        else: 
            print(f'no hay registros nuevos') 
        print('----------------------------------------')
        
    def registrar_seven(self,registro:list):  
        self.conexion.execute("INSERT INTO pure VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ? ,?)", (str(registro[1]) , str(registro[2]) , str(registro[3]) , str(registro[4]) ,str(registro[5]) , str(registro[6]),str(registro[7]) , str(registro[8]) , int(registro[9]) , int(registro[10]) ,int(registro[11]) , str(registro[12]), str(registro[13]) , str(registro[14])))
  

# seven = Archivo_seven('/gastos_seven.xlsx')
# print(len(seven.registros))
# seven.imprimir_seven()

# pure = Archivo_pure('/gastos_pure.xlsx') 
# pure.imprimir_pure()


# Proy=archivo_proyecto('/proyectos.xlsx')
# Proy.imprimir_proyectos()
# seven.print()





if __name__ == "__main__":


    def menu():
        print('            Migrar de SEVEN a PURE:')
        print('Menú')
        print('')
        print('Cargar archivos : (1)')
        print('Salir (0)')
        print('----------------------------------------')

    rta = 9999
    menu()

    while rta != 0:

        rta = input('seleccione una del las opciones anteriores: ')
        if rta.isnumeric():
            rta = int(rta)
            if 0 <= rta and rta <= 2 :
                if rta == 0:
                    break

                elif rta == 1 :
                    seven = Archivo_seven('/gastos_seven.xlsx')
                    pure = Archivo_pure('/gastos_pure.xlsx') 
                    
                    if seven.wb != None and pure.wb != None and len(seven.proyectos.proyectos)>0:
                        rta2 = 9999
                        while rta2 != 0:
                            err = len(seven.registros_error)
                            if err > 0:
                                print('\n\n')
                                print('Existen '+ str(err) +' errores' )
                                print('Desea continar  \n (1) Registrar en pure \n (2) Exportar Pure \n (3) Imprimir errores \n (0) regresar')
                                print('----------------------------------------')
                                rta2=input('seleccione una del las opciones anteriores: ')
                                if rta2.isnumeric():
                                    rta2 = int(rta2)
                                    if 0 <= rta2 and rta2 <= 4 :
                                        if rta2 == 1:
                                            pure.verificar_registros(seven)
                                        elif rta2 == 2:
                                            pass
                                        elif rta2 == 3:
                                            seven.imprimir_errores()

                                        elif rta2 ==0:
                                            break
                                    else:
                                        print('\n\n----------------------------------------')     
                                        print('                           opcion no valida, no esta dentro del rango')
                                else:
                                    print('\n\n----------------------------------------')  
                                    print('                             opcion no valida, no es un entero')     
                    else:
                        print('\n\n----------------------------------------')
                        print('verifique que el archivo existe o tenga el nombre correcto')    

            else:
                print('\n\n----------------------------------------')     
                print('                           opcion no valida, no esta dentro del rango')
        else:
            print('\n\n----------------------------------------')  
            print('                             opcion no valida, no es un entero')      

        rta = 999
        print('----------------------------------------')   
        print('')  
        menu()

    print('cerró app: ')    